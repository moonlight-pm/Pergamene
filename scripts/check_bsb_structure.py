#!/usr/bin/env python3
"""
Check the structure of the BSB Excel file
"""

from openpyxl import load_workbook
import sys

excel_path = sys.argv[1] if len(sys.argv) > 1 else "downloads/bsb.xlsx"

print(f"Loading {excel_path}...")
wb = load_workbook(excel_path, read_only=True)

print(f"Sheet names: {wb.sheetnames}")

for sheet_name in wb.sheetnames:
    sheet = wb[sheet_name]
    print(f"\n=== Sheet: {sheet_name} ===")
    
    # Get first 5 rows to understand structure
    rows = []
    for i, row in enumerate(sheet.iter_rows(values_only=True)):
        if i >= 5:
            break
        rows.append(row)
    
    # Print the rows
    for i, row in enumerate(rows):
        print(f"Row {i}: {row[:6] if len(row) > 6 else row}")  # First 6 columns
    
    # Count total rows
    row_count = sum(1 for _ in sheet.iter_rows())
    print(f"Total rows in {sheet_name}: {row_count}")

wb.close()