#!/usr/bin/env python3
"""
Parse BSB Excel file with correct structure
"""

import json
import sys
import re
from pathlib import Path
from openpyxl import load_workbook

def parse_reference(ref_text):
    """Parse a reference like 'Genesis 1:1' into book, chapter, verse"""
    match = re.match(r'^(.+?)\s+(\d+):(\d+)$', ref_text.strip())
    if match:
        return match.group(1), int(match.group(2)), int(match.group(3))
    return None, None, None

def parse_bsb_nt(excel_path, output_path):
    """Parse BSB Excel file and extract NT books"""
    
    print(f"Reading BSB Excel file from {excel_path}...")
    wb = load_workbook(excel_path, read_only=True)
    sheet = wb['Sheet1']
    
    # NT books
    nt_books = [
        'Matthew', 'Mark', 'Luke', 'John', 'Acts',
        'Romans', '1 Corinthians', '2 Corinthians', 'Galatians', 'Ephesians',
        'Philippians', 'Colossians', '1 Thessalonians', '2 Thessalonians',
        '1 Timothy', '2 Timothy', 'Titus', 'Philemon',
        'Hebrews', 'James', '1 Peter', '2 Peter',
        '1 John', '2 John', '3 John', 'Jude', 'Revelation'
    ]
    
    # Create a set for faster lookup
    nt_books_set = set(nt_books)
    
    # Parse the data
    books = {}
    verse_count = 0
    nt_verse_count = 0
    
    for row_num, row in enumerate(sheet.iter_rows(values_only=True)):
        # Skip header rows
        if row_num < 3:
            continue
            
        if len(row) < 3:
            continue
        
        # Column B has reference, Column C has text
        reference = row[1]
        text = row[2]
        
        if not reference or not text:
            continue
        
        book_name, chapter, verse = parse_reference(str(reference))
        
        if not all([book_name, chapter, verse]):
            continue
        
        verse_count += 1
        
        # Check if this is an NT book
        if book_name not in nt_books_set:
            continue
            
        nt_verse_count += 1
        
        # Initialize book if needed
        if book_name not in books:
            books[book_name] = {
                'name': book_name,
                'testament': 'New',
                'chapters': {}
            }
        
        # Initialize chapter if needed
        chapter_str = str(chapter)
        if chapter_str not in books[book_name]['chapters']:
            books[book_name]['chapters'][chapter_str] = []
        
        # Add verse
        books[book_name]['chapters'][chapter_str].append({
            'number': verse,
            'text': str(text).strip()
        })
    
    wb.close()
    
    print(f"Processed {verse_count} total verses")
    print(f"Found {nt_verse_count} NT verses")
    
    # Convert to ordered structure matching OT format
    ordered_books = []
    for book_name in nt_books:
        if book_name in books:
            book_data = books[book_name]
            # Sort chapters numerically
            sorted_chapters = []
            for ch_num in sorted(book_data['chapters'].keys(), key=int):
                sorted_chapters.append({
                    'number': int(ch_num),
                    'verses': sorted(book_data['chapters'][ch_num], key=lambda v: v['number'])
                })
            
            ordered_books.append({
                'name': book_name,
                'testament': 'New',
                'chapters': sorted_chapters
            })
    
    # Save to JSON
    print(f"Writing NT data to {output_path}...")
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(ordered_books, f, indent=2, ensure_ascii=False)
    
    print(f"Successfully extracted {len(ordered_books)} NT books")
    
    # Print summary
    for book in ordered_books:
        num_chapters = len(book['chapters'])
        num_verses = sum(len(ch['verses']) for ch in book['chapters'])
        print(f"  {book['name']}: {num_chapters} chapters, {num_verses} verses")

if __name__ == '__main__':
    if len(sys.argv) != 3:
        print("Usage: python parse_bsb_nt_correct.py <excel_file> <output_json>")
        sys.exit(1)
    
    excel_path = Path(sys.argv[1])
    output_path = Path(sys.argv[2])
    
    if not excel_path.exists():
        print(f"Error: Excel file not found: {excel_path}")
        sys.exit(1)
    
    try:
        parse_bsb_nt(excel_path, output_path)
    except Exception as e:
        print(f"Error parsing BSB: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)