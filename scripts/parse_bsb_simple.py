#!/usr/bin/env python3
"""
Simple BSB Excel parser using openpyxl (no pandas required)
"""

import json
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("Error: openpyxl not installed")
    print("Run: pip3 install openpyxl")
    sys.exit(1)

def parse_bsb_nt(excel_path, output_path):
    """Parse BSB Excel file and extract NT books"""
    
    print(f"Reading BSB Excel file from {excel_path}...")
    wb = load_workbook(excel_path, read_only=True)
    
    # Find the BSB sheet
    if 'BSB' in wb.sheetnames:
        sheet = wb['BSB']
    else:
        # Try the first sheet
        sheet = wb.active
    
    # NT books
    nt_books = [
        'Matthew', 'Mark', 'Luke', 'John', 'Acts',
        'Romans', '1 Corinthians', '2 Corinthians', 'Galatians', 'Ephesians',
        'Philippians', 'Colossians', '1 Thessalonians', '2 Thessalonians',
        '1 Timothy', '2 Timothy', 'Titus', 'Philemon',
        'Hebrews', 'James', '1 Peter', '2 Peter',
        '1 John', '2 John', '3 John', 'Jude', 'Revelation'
    ]
    
    # Create a set for faster lookup
    nt_books_set = set(nt_books)
    
    # Parse the data
    books = {}
    header_row = True
    verse_count = 0
    nt_verse_count = 0
    
    for row in sheet.iter_rows(values_only=True):
        if header_row:
            header_row = False
            continue
        
        if len(row) < 4:
            continue
            
        book_name = row[0]
        chapter = row[1]
        verse = row[2]
        text = row[3]
        
        if not all([book_name, chapter, verse, text]):
            continue
        
        verse_count += 1
        
        # Check if this is an NT book
        if book_name not in nt_books_set:
            continue
            
        nt_verse_count += 1
        
        # Initialize book if needed
        if book_name not in books:
            books[book_name] = {
                'name': book_name,
                'testament': 'New',
                'chapters': {}
            }
        
        # Initialize chapter if needed
        chapter_str = str(chapter)
        if chapter_str not in books[book_name]['chapters']:
            books[book_name]['chapters'][chapter_str] = []
        
        # Add verse
        books[book_name]['chapters'][chapter_str].append({
            'number': int(verse),
            'text': str(text).strip()
        })
    
    wb.close()
    
    print(f"Processed {verse_count} total verses")
    print(f"Found {nt_verse_count} NT verses")
    
    # Convert to ordered structure
    ordered_books = []
    for book_name in nt_books:
        if book_name in books:
            book_data = books[book_name]
            # Sort chapters numerically
            sorted_chapters = []
            for ch_num in sorted(book_data['chapters'].keys(), key=int):
                sorted_chapters.append({
                    'number': int(ch_num),
                    'verses': sorted(book_data['chapters'][ch_num], key=lambda v: v['number'])
                })
            
            ordered_books.append({
                'name': book_name,
                'testament': 'New',
                'chapters': sorted_chapters
            })
    
    # Save to JSON
    print(f"Writing NT data to {output_path}...")
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(ordered_books, f, indent=2, ensure_ascii=False)
    
    print(f"Successfully extracted {len(ordered_books)} NT books")
    
    # Print summary
    for book in ordered_books:
        num_chapters = len(book['chapters'])
        num_verses = sum(len(ch['verses']) for ch in book['chapters'])
        print(f"  {book['name']}: {num_chapters} chapters, {num_verses} verses")

if __name__ == '__main__':
    if len(sys.argv) != 3:
        print("Usage: python parse_bsb_simple.py <excel_file> <output_json>")
        sys.exit(1)
    
    excel_path = Path(sys.argv[1])
    output_path = Path(sys.argv[2])
    
    if not excel_path.exists():
        print(f"Error: Excel file not found: {excel_path}")
        sys.exit(1)
    
    try:
        parse_bsb_nt(excel_path, output_path)
    except Exception as e:
        print(f"Error parsing BSB: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)